from bs4 import BeautifulSoup
from requests_html import HTMLSession
import pandas as pd
import urllib.parse
import openpyxl
from tkinter import messagebox as mb
import os
import os.path as op
def youtube_data(url):
    session=HTMLSession()
    res=session.get(url)
    res.html.render(sleep=8,timeout=1000)
    soup=BeautifulSoup(res.html.html,"html.parser")
    videos_data=soup.findAll("ytd-video-renderer")
    data_list=[]
    for data in videos_data:
        titles=data.find("a",{"class":"yt-simple-endpoint style-scope ytd-video-renderer"}).text
        urls=urllib.parse.urljoin("https://www.youtube.com",data.find("a",{"class":"yt-simple-endpoint style-scope ytd-video-renderer"}).get("href"))
        channel=data.find("a",{"class":"yt-simple-endpoint style-scope yt-formatted-string"})
        channel_name=channel.text
        channel_url=urllib.parse.urljoin("https://www.youtube.com",channel.get("href"))
        description=data.find("yt-formatted-string",{"class":"metadata-snippet-text style-scope ytd-video-renderer"})
        try:
            try:
                description1=description.find_next("span")
            except:
                pass
            try:
                description2=description.find_next("span").find_next("span")
            except:
                pass
            try:
                description3=description.find_next("span").find_next("span").find_next("span")
            except:
                pass
            try:
                description4=description.find_next("span").find_next("span").find_next("span").find_next("span")
            except:
                pass

        except:
            pass
        DESCRIPTIONS=f"{description1.text}{description2.text}{description3.text}{description4.text}"
        meta_data=data.find("div",id="metadata")
        views=meta_data.find_next("span").text
        publ_date=meta_data.find_next("span").find_next("span").next
        data_results={
            "Title":titles,
            "Views":views,
            #"Date":publ_date,
            "Description":DESCRIPTIONS,
            "Url":urls,
            "Channel Name":channel_name,
            "Channel Url":channel_url
        }
        data_list.append(data_results)
    return data_list
def results(data_list):
    dataResult=pd.DataFrame(data_list)
    dataResult.to_csv("youtube-data.csv",index=False)
    print("Done")
    


def data_extraction():
    search=y.get()
    data=youtube_data(f"https://www.youtube.com/results?search_query={search}")
    Result=results(data)
    if op.exists("youtube-data.csv"):
        mb.showinfo("Warnings ","Data extraction is successful.")

    data1 = pd.read_csv("youtube-data.csv")
    wb = openpyxl.Workbook()
    sayfa = wb.active

    a2 = len(data1)           ### toplan satır sayısı
    a3 = len(data1.columns)   ### toplam sütun sayısı
    print('satır uzunluğu: ', a2)
    print('sütun sayısı: ', a3)


    for x in range(a3):      ### sütun başlıklarını yazdırma döngüsü
        c = x + 1
        sayfa.cell(row = 1, column = c).value = data1.columns[x]


    for x in range(a2):    ### tüm satırlardaki verileri excele yazdırma döngüsü
        for s in range(a3):
            r = x + 2
            c = s + 1
            sayfa.cell(row = r, column = c).value = data1.iat[x,s]


    wb.save("youtube-data-excell.xlsx")
    os.remove("youtube-data.csv")
    print('İşlem başarıyla tamamlandı. Excel dosyanız oluşturuldu')
    if op.exists("youtube-data-excell.xlsx"):
        mb.showinfo("Warnings ","You can access the excel file in the downloaded folder.")
import tkinter as tk
arayüz = tk.Tk()
arayüz.configure(background="black")
arayüz.title("Youtube Search Results Analyse ")
arayüz.geometry("600x500")
arayüz.resizable(width=False,height=False)

etiket = tk.Label(text="ENTER SUBJECT:",background="silver")
etiket.place(x=200,y=300)

y = tk.StringVar()
search = tk.Entry(textvariable=y,background="White Smoke")
search.place(x=300, y=300)

tahmin = tk.Button(text="GUESS")
tahmin.place(x=200, y=340)

veri = tk.Button(text="DATA EXTRACTION",command=data_extraction,activebackground="Plum")
veri.place(x=280,y=340)

#excell = tk.Button(text="Excell View",command=csvExeceleDonus)
#excell.place(x=200, y=380)
arayüz.mainloop()